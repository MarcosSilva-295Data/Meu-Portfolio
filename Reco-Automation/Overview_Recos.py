import pandas as Pd
import  openpyxl as op
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
import funções 
from datetime import datetime
from dateutil.relativedelta import relativedelta
import locale

locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

def criar_pivot(status_nome, df):
    tiers = ["Reg.", "T0", "T1", "T2", "T3"]
    pivot = Pd.pivot_table(
        df[df["Status"] == status_nome],
        index="Area de Atendimento",
        columns="Tier",
        values="Code",
        aggfunc="count",
        fill_value=0
    ).reindex(columns=tiers, fill_value=0)

    pivot["Total"] = pivot.sum(axis=1)
    pivot.columns = Pd.MultiIndex.from_product([[status_nome], pivot.columns])

    return pivot


def largura_col(ws):
    for col in ws.columns:
        valores = [len(str(c.value)) for c in col if c.value]

        largura = max(valores) if valores else 1
        largura = max(5, min(largura, 12))

        ws.column_dimensions[col[0].column_letter].width = largura

def Overview_forun(overview, df1):
    
    df = df1[df1['Re_Status'].isin(['Open', 'In Committee Review'])].copy()

    df['Status'] = df['Status Vencimento'].map({
        'A vencer em 30 dias': 'NO PRAZO',
        'A vencer em 60 dias': 'NO PRAZO',
        'A vencer em 90 dias': 'NO PRAZO',
        'A vencer em mais de 90 dias': 'NO PRAZO',
        'Em atraso': 'EM ATRASO (sem evidência)',
        'Análise em evidência': 'EM ANÁLISE'
    })

    p1 = criar_pivot("NO PRAZO", df)
    p2 = criar_pivot("EM ANÁLISE", df)
    p3 = criar_pivot("EM ATRASO (sem evidência)", df)

    tabela_final = Pd.concat([p1, p2, p3], axis=1).fillna(0)

    tabela_final[("TOTAL GERAL", "Total")] = (
        tabela_final[("NO PRAZO", "Total")] +
        tabela_final[("EM ANÁLISE", "Total")] +
        tabela_final[("EM ATRASO (sem evidência)", "Total")]
    )

    tabela_final = tabela_final.sort_values(by=("TOTAL GERAL", "Total"), ascending=False)

    # remover colunas zeradas
    tabela_final = tabela_final[
        [col for col in tabela_final.columns if col[1] == "Total" or (tabela_final[col] != 0).any()]
    ]

    # ---------------------------
    # TOTAL FINAL
    # ---------------------------
    total_linha = tabela_final.sum().to_frame().T
    total_linha.index = ["Total"]
    tabela_final = Pd.concat([tabela_final, total_linha])

    # ---------------------------
    # EXPORTAÇÃO
    # ---------------------------
    with Pd.ExcelWriter(overview, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        tabela_final.to_excel(writer, sheet_name=f'Overview{datetime.now().strftime("%B").capitalize()}', startrow=1)

    wb = op.load_workbook(overview)
    ws = wb[f'Overview{datetime.now().strftime("%B").capitalize()}']

    ws.delete_rows(4)

    largura_col(ws)

    max_linha = ws.max_row
    max_coluna = ws.max_column

    # ---------------------------
    # TÍTULOS
    # ---------------------------
    ws['A1'] = 'Estoque de Recomendações'
    ws['A2'] = 'Área Respónsavel Pelo Atendimento'

    # ---------------------------
    # ESTILO BASE
    # ---------------------------
    font_padrao = Font(size=9)
    align_centro = Alignment(horizontal='center', vertical='center')

    cores_header = {
        "NO PRAZO": ('4EA72E'),
        "EM ANÁLISE": ('44B3E1'),
        "EM ATRASO (sem evidência)": ('C00000'),
        "TOTAL GERAL": ('156082')
    }

    # ---------------------------
    # FORMATAÇÃO GERAL
    # ---------------------------
    for row in ws.iter_rows(min_row=1, max_row=max_linha, max_col=max_coluna):
        for cell in row:
            cell.font = font_padrao
            cell.alignment = align_centro

            if cell.value in cores_header:
                cell.font = Font(bold=True, size=9, color='FFFFFF')
                cell.fill = PatternFill("solid", fgColor=cores_header[cell.value])

    # ---------------------------

    # ---------------------------
    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=max_linha, max_col=max_coluna)):
        cor = 'FFFFFF' if i % 2 == 0 else 'E8E8E8'
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=cor)

    # ---------------------------
    # HEADER COLORIDO (tiers)
    # ---------------------------
    cores_tier = ['DAF2D0', 'DAE9F8', 'FF8E85']
    celulas_totais = []
    cont = 0

    for col_idx, cell in enumerate(ws[3][1:-1]):
        cell.fill = PatternFill("solid", fgColor=cores_tier[cont])

        if cell.value == 'Total':
            celulas_totais.append(col_idx)
            cont += 1

    # ---------------------------
    # NEGRITO TOTAL
    # ---------------------------
    for cell in ws.iter_rows(min_row=4, max_row=max_linha, min_col=max_coluna, max_col=max_coluna):
        for c in cell:
            c.font = Font(bold=True, size=9)

    for cell in ws.iter_rows(min_row=max_linha, max_row=max_linha, min_col=2, max_col=max_coluna):
        for c in cell:
            c.font = Font(bold=True, size=9)

    # ---------------------------
    # COLORIR TOTAIS POR BLOCO
    # ---------------------------
    cores = ['4EA72E', '44B3E1', 'C00000']

    for row in ws.iter_rows(min_row=3, max_row=max_linha, min_col=2, max_col=max_coluna - 1):
        cont = 0
        for i, cell in enumerate(row):
            if i in celulas_totais:
                cell.font = Font(bold=True, size=9, color=cores[cont])
                cont += 1

    # ---------------------------
    # MERGES
    # ---------------------------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_coluna)
    ws['A1'].font = Font(bold=True, size=9, color='FFFFFF')
    ws['A1'].fill = PatternFill("solid", fgColor='535B5B')

    ws.merge_cells('A2:A3')
    ws['A2'].font = Font(bold=True, size=9)
    ws['A2'].fill = PatternFill("solid", fgColor='E8E8E8')
      
    
    ws.merge_cells(start_row=2, start_column=max_coluna, end_row=3, end_column=max_coluna)
    gerar_extrato(df1,ws,max_coluna)

    prox_venc = proximos_vencimento(df)
    
    ws.cell(row=max_linha+2, column=1, value='Próximos Vencimentos').fill = PatternFill("solid", fgColor='535B5B')
    ws.cell(row=max_linha+2, column=1).font = Font(bold=True, size=9, color='FFFFFF')
    ws.cell(row=max_linha+2, column=1).alignment = op.styles.Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=max_linha+2, start_column=1, end_row=max_linha+2, end_column=len(prox_venc.columns)+1)
    wb.save(overview)
    
    with Pd.ExcelWriter(overview, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        prox_venc.to_excel(writer, sheet_name=f'Overview{datetime.now().strftime("%B").capitalize()}', startrow=max_linha+2)    
    
    
        
def Metricas(df1, local):
    base_recos = df1[df1['Re_Status'].isin(['Open','In Committee Review'])]
    wb = op.load_workbook(r"Template.xlsx")
    wb.worksheets[0].title = f'BASE_RECOS'
    
    e = wb['BASE_RECOS']
    x=0
    for row in e.iter_rows(min_row=1, max_row=len(base_recos.index)+1, max_col=len(base_recos.columns)):
        y=0
        if x==0:
            for cell in row:
                cell.value = base_recos.columns[y]
                cell.font = op.styles.Font(bold=True)
                cell.fill= op.styles.PatternFill(fill_type='solid', fgColor='A6A6A6')
                cell.alignment = op.styles.Alignment(horizontal='center', vertical='center')
                y+=1 
        else:
            linha= base_recos.iloc[x-1]
            for cell in row:
                cell.value = linha.iloc[y]
                y+=1     
        x+=1 
    colunas_com_datas = ["H","AA","AC","AJ","AB","AD","AE","AF","AN","AZ","BN","BO","BQ"]       
    funções.largura_col(e)
    for i in colunas_com_datas:
             funções.Ajuste_data(i, e)
             
    wb.create_sheet(title=f'Metricas')
    e = wb[f'Metricas']
    
    df_ant= df1[df1['Re_Status Anterior'].isin(['Open','In Committee Review'])]
    df = df1[df1['Re_Status'].isin(['Open','In Committee Review'])]
    Colunas=['B','C','D','E','F','G','H']
    divisoes = ["Regulátorias", "Tier 1", "Tier 2 e 3","Total"]
    reg_ant = df_ant[df_ant['Tier'].isin(['Reg.'])]
    reg = df[df['Tier'].isin(['Reg.'])]
    t1_ant = df_ant[df_ant['Tier'].isin(['T1','T0'])]
    t1 = df[df['Tier'].isin(['T1','T0'])]  
    t2e3_ant = df_ant[df_ant['Tier'].isin(['T2','T3'])]
    t2e3 = df[df['Tier'].isin(['T2','T3'])]  
    planilhas =[(reg, reg_ant),(t1, t1_ant),(t2e3,t2e3_ant),(df, df_ant)] 
    lin =0
    cont=0
    for linha in range(2,8):
        col=0

        x=planilhas[cont][0]
        y=planilhas[cont][1]
        atrasadas_ant = len(y[y['Status/Venc anterior'].isin(['Em atraso'])])
        atrasadas = len(x[x['Status Vencimento'].isin(['Em atraso'])])   
        Linhas =[("Tier", f'{(datetime.today()- relativedelta(months=2)).strftime("%B")}', ""," ", f'{(datetime.today()- relativedelta(months=1)).strftime("%B")}',"", " "),
                (" ","Ativas", "Vencidas", 'cálculo %',"Ativas" ,"Vencidas", 'cálculo %'),
                (divisoes[cont],len(y['Re_Status Anterior']), atrasadas_ant, f'{round((atrasadas_ant/len(y['Re_Status Anterior']))*100,1)}%', len(x['Re_Status']),
                 atrasadas, f'{round((atrasadas/len(x['Re_Status']))*100,1)}%')]
               
        for i in Colunas:
            
            e[f'{i}{linha}'].value = Linhas[lin][col]
            col+=1
            
        if lin >=2: 
            cont+=1
        else: lin+=1   
    
    e.merge_cells('C2:E2')
    e.merge_cells('F2:H2')
    e.merge_cells('B2:B3')
    
    for row in e['B2:H7']:
        for cell in row:
            cell.alignment = op.styles.Alignment(horizontal='center', vertical='center')
            cell.border = op.styles.Border(left=op.styles.Side(style='dotted'),
                                            right=op.styles.Side(style='dotted'), 
                                            top=op.styles.Side(style='dotted'), 
                                            bottom=op.styles.Side(style='dotted'))
    wb.save(local+f'Metricas_{datetime.now().strftime("%B").capitalize()}.xlsx')
    
    
    

def gerar_extrato(df,planilha,max_coluna):

    dta_ref =df['Extraction Date'].iloc[0] + Pd.DateOffset(months=1)
    dta_ref_ant=df['Extraction Date'].iloc[0] + Pd.DateOffset(months=1)
    # 🔹 TOTAL GERAL
    total_geral = {
        "Total anterior": df['Re_Status Anterior'].isin(['Open', 'In Committee Review']).sum(),
        "Canceladas": -(df['Alterações do fechamento'] == 'Cancelada').sum(),
        "Atendidas": -(df['Alterações do fechamento'] == 'Atendida').sum(),
        "Novas emissões": (df['Alterações do fechamento'] == 'Nova Reco').sum(),
        "Total Atual": df['Re_Status'].isin(['Open', 'In Committee Review']).sum()
    }
    df_total = Pd.DataFrame.from_dict(total_geral, orient='index',columns=["Total Geral"])

    # 🔹 EM ATRASO
    filtro_atraso_ant = df['Status/Venc anterior'] == 'Em atraso'

    em_atraso = {
        "Total anterior": filtro_atraso_ant.sum(),

        "Canceladas": -(filtro_atraso_ant &
                       (df['Alterações do fechamento'] == 'Cancelada')).sum(),

        "Atendidas": -(filtro_atraso_ant &
                      (df['Alterações do fechamento'] == 'Atendida')).sum(),

        "Recalendarizadas": -(filtro_atraso_ant &
                             df['Recalendarizada'].isin(['1ª', '+1'])).sum(),

        "Recebeu evidência": -(filtro_atraso_ant &
                              (df['Alterações do fechamento'] == 'Recebeu Evidência')).sum(),

        "Evidência recusada": (filtro_atraso_ant &
                               (df['Alterações do fechamento'] == 'Evidência recusada')).sum(),

        "Novo vencimento": ((df['Status/Venc anterior'] != 'Em atraso') &
                            (df['Status Vencimento'] == 'Em atraso')).sum(),

        "Total Atual": (df['Status Vencimento'] == 'Em atraso').sum()
    }
    df_atraso = Pd.DataFrame.from_dict(em_atraso, orient='index',columns=["Em Atraso"])
    # =========================
    # 🔹 EM ANÁLISE
    # =========================
    filtro_analise_ant = df['Status/Venc anterior'] == 'Análise em evidência'

    em_analise = {
        "Total anterior": filtro_analise_ant.sum(),

        "Canceladas": -(filtro_analise_ant &
                       (df['Alterações do fechamento'] == 'Cancelada')).sum(),

        "Atendidas": -(filtro_analise_ant &
                      (df['Alterações do fechamento'] == 'Atendida')).sum(),

        "Evidência Retornou": -(filtro_analise_ant &
                               (df['Alterações do fechamento'] == 'Evidência recusada')).sum(),

        "Nova Evidência": 
                           (df['Alterações do fechamento'] == 'Recebeu Evidência').sum(),

        "Total Atual": (df['Status Vencimento'] == 'Análise em evidência').sum()
    }
    df_analise = Pd.DataFrame.from_dict(em_analise, orient='index',columns=["Em análise"])
    # =========================
    # 🔹 MONTA TABELA FINAL
    blocos= [('Total Geral', total_geral),('Em Atraso', em_atraso),('Em análise',em_analise)]
    cores_tier = ['156082', 'C00000', '44B3E1']
    col_inicio = max_coluna+ 2
    planilha.cell(row=1, column=col_inicio, value='Extrato').fill = PatternFill("solid", fgColor='535B5B')
    planilha.cell(row=1, column=col_inicio).font = Font(bold=True, size=9, color='FFFFFF')
    planilha.cell(row=1, column=col_inicio).alignment = op.styles.Alignment(horizontal='center', vertical='center')
    planilha.merge_cells(start_row=1, start_column=col_inicio, end_row=1, end_column=col_inicio+5)
    
    
    for i, (titulo, dados)in enumerate(blocos):
        col=col_inicio+(i*2)
        planilha.cell(row=2, column=col, value=titulo).font = Font(bold=True, size=9, color='FFFFFF')
        planilha.cell(row=2, column=col, value=titulo).fill = PatternFill("solid", fgColor=cores_tier[i])
        planilha.cell(row=2, column=col, value=titulo).alignment = op.styles.Alignment(horizontal='center', vertical='center')
        planilha.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
        linha=3
        for k,v in dados.items():
            planilha.cell(row=linha, column=col, value=k).font = Font(bold=True,size=9)
            planilha.cell(row=linha, column=col).alignment = op.styles.Alignment(horizontal='center', vertical='center')
            planilha.cell(row=linha, column=col+1, value=v).font = Font(size=9)
            planilha.cell(row=linha, column=col+1).alignment = op.styles.Alignment(horizontal='center', vertical='center')
            linha+=1


def classificar_periodo(data, data_ref):
    if Pd.isna(data):
        return None

    tri_ref = (data_ref.month - 1) // 3 + 1
    tri_data = (data.month - 1) // 3 + 1

    ano_ref = data_ref.year
    ano_data = data.year

    # -------------------------
    # MESMO ANO
    # -------------------------
    if ano_data == ano_ref:

        if tri_ref == 1:
            return f"{tri_data}ºTri"

        elif tri_ref == 2:
            if tri_data >= 2:
                return f"{tri_data}ºTri"

        elif tri_ref == 3:
            if tri_data >= 3:
                return f"{tri_data}ºTri"

        elif tri_ref == 4:
            if tri_data == 4:
                return "4ºTri"

        return None # evita cair em outras regras

    # -------------------------
    # PRÓXIMO ANO (Q4 especial)
    # -------------------------
    if ano_data != ano_ref and ano_data == ano_ref+1 :
        if tri_ref == 4 :
            return f"{tri_data}ºTri_{ano_data}"
        else:
            return str(ano_data)
    else:
        return str(ano_data)

def ordenar_periodo(col):
    
    if col == "Total":
        return(99,0)
    # Trimestres do ano atual (ex: 1ºTri)
    if "Tri" in col and "_" not in col:
        num = int(col[0])
        return (0, num)

    # Trimestres com ano (ex: 1ºTri_2027)
    if "Tri_" in col:
        tri = int(col[0])
        ano = int(col.split("_")[1])
        return (1, ano, tri)

    # Anos (ex: 2027, 2028)
    return (2, int(col))
    

def proximos_vencimento(df1):
    dta_ref =df1['Extraction Date'].iloc[0] + Pd.DateOffset(months=1)
    df = df1[~(df1['Status Vencimento'].isin(['Em atraso','Análise em evidência']))]
   
    df['periodo']= df['Current due Date'].apply(lambda x: classificar_periodo(x,dta_ref))
    
    pivot = Pd.pivot_table(
        df,
        index="Area de Atendimento",
        columns="periodo",
        values="RE Code",
        aggfunc="count",
        margins=True,
        margins_name='Total',
        fill_value="-"
    )
    pivot = pivot[sorted(pivot.columns, key=ordenar_periodo)]
    return pivot
    



