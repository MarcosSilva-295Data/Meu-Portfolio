import pandas as Pd
import openpyxl as op
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import Overview_Recos

def Status_Vencimento(dias, status, restatus):
    if restatus in ["Cancelled", "Closed", "Deleted"] or Pd.isna(status):
        return "Não ativo"
    elif status=='In Evidence Review':
        return "Análise em evidência"
    elif dias < 0:
        return "Em atraso"
    elif dias <= 30 :
        return "A vencer em 30 dias"
    elif dias <= 60 :
        return "A vencer em 60 dias"
    elif dias <= 90 :
        return "A vencer em 90 dias"
    else:
        return "A vencer em mais de 90 dias"

def area_responsavel(manager, segmento):
    equipe_atendimento = {
        "Risk Adequacy":['Team_Metodologia','Team_Provisao','Team_Stress Test',
                                          'Team_IRB','Team_Riesgos','Team_CRC'],
        "Market Risk":['Mercado e Contraparte','Risco Estrutural'],
        "Risk Analytics":['Team_PM_PJ_Agro','Team_Varejo_PF','Team_Motores', 'Team_Methd'],
        "SCIB Risk":['Team__PM SCIB e Private'],
        "Plataforma Risk":['Team_Projetos Bdr','Team_Plataforma Informacional'],
        "Capital Risk":['Team_Capital'],
        "Juridico Risk":['Team_Juridico'],
        "Operacional Risk":['Team_Risco Operacional']
    }
    if manager in equipe_atendimento['Risk Adequacy'] and segmento not in equipe_atendimento['Market Risk']:
        return "Risk Adequacy"
    if segmento in equipe_atendimento['Market Risk']:
        return "Market Risk"
    for area, lista in equipe_atendimento.items():
        if manager in lista:
            return area
        
    return manager
    
def AlteracoesFechamento(atual,anterior,Fianterior, FiAtual):
    
    if anterior in ['Open','In Committee Review'] and atual =='Closed':
        return"Atendida"
    
    elif (anterior in ['Open','In Committee Review']) and (atual =='Cancelled' or atual == 'Deleted'):
        return"Cancelada"
    
    elif Fianterior=='In Evidence Review' and FiAtual =='In Evidence Submission':
        return 'Evidência recusada'
    
    elif FiAtual=='In Evidence Review' and Fianterior =='In Evidence Submission':
        return 'Recebeu Evidência'
    
    elif atual=='Open' and Pd.isna(anterior):
        return "Nova Reco"
    
    elif (atual==anterior and Fianterior==FiAtual) or (anterior in ['Open','In Committee Review'] and atual in ['Open','In Committee Review']): 
        return'Sem alteração'
    
    else:
        return'-'

def Recalendarizacao(anterior, atual, inicial):
    if Pd.isna(anterior):
        return "Não"
    elif anterior != atual and anterior == inicial:
        return "1ª"
    elif anterior !=atual and anterior != inicial:
        return "+1"
    elif anterior == atual and atual != inicial:
        return "Pelo menos uma"
    else:
        return "Não"

def largura_col(planilha):
    for col in planilha.columns:
        
        largura = [len(str(celula.value)) for celula in col if celula.value]
        
        if largura:
            largura = max(largura)
        else: largura = 1
        if largura >30:
            largura = 30
        if largura < 10:
            largura = 10
        coluna_letra=col[0].column_letter
        planilha.column_dimensions[coluna_letra].width = largura

def Ajuste_data(Coluna, planilha):
      for row in range(2,planilha.max_row + 1):
            cell = planilha[f"{Coluna}{row}"]
            cell.number_format = "DD/MM/YYYY"

def Base(Finding_base,Finding_Comparativa,inventario,recommendation, local):
    
    #Lê a Planilhas que serão utilizadas para gerar o estoque
    df_base = Pd.read_excel(Finding_base)
    df_anterior = Pd.read_excel(Finding_Comparativa)
    df_Recommendation = Pd.read_excel(recommendation)
    df_inventario = Pd.read_excel(inventario, header=1)
    
    #Aplica o filtro na coluna Responsible Legal Entity  no Data Frame da tabela base
    df_base = df_base[(df_base['Responsible Legal Entity'].isin(['Team_Brasil']))]
    
    #Pega a data de extração das tabelas base e anterior e soma mais um mês
    dta_ref =df_base['Extraction Date'].iloc[0] + Pd.DateOffset(months=1)
    dta_ref_ant=df_anterior['Extraction Date'].iloc[0] + Pd.DateOffset(months=1)

    #Lê a Planilha Recommendation e pega a coluna Number of Re-Schedules e passa para a Base usando o Code
    df_Recommendation =df_Recommendation[['Code','Number of Re-Schedules']]
    df_Recommendation.rename(columns={'Number of Re-Schedules':'Qntd de Recalendarizações'},inplace=True)
    df_base = Pd.merge(df_base, df_Recommendation, on = 'Code', how = 'left')
    
    #Lê a Planilha do Invetario e pega a coluna Segmentação Brasil e passa para Base usando o Code
    df_inventario =df_inventario[['Code','Segmentação Brasil']]
    df_inventario.rename(columns={'Segmentação Brasil':'Segmento Brasil'},inplace=True)
    df_base = Pd.merge(df_base, df_inventario, on = 'Code', how = 'left')
    
    #Cria na Planilha anterior as colunas Dias Até Atraso e Status Vencimento
    df_anterior['Current due Date'] = Pd.to_datetime(df_anterior['Current due Date'], errors='coerce')
    df_anterior['Dias Até Atraso'] = (df_anterior['Current due Date'] - dta_ref_ant).dt.days
    df_anterior['Status Vencimento'] = df_anterior.apply(lambda row:Status_Vencimento(row['Dias Até Atraso'], row['Status Reco Finding'], row['Reco Status']), axis=1)
    
    #Passa as colunas Reco Status, Status Reco Finding e Status de vencimento da Planilha Finding anterior para a Planilha Base
    df_anterior = df_anterior[['FI_Code','Current due Date','Re_Status','Status Reco Finding','Status Vencimento']]
    df_anterior.rename(columns={'Current due Date':'Data de Vencimento Anterior',
                                'Re_Status':'Re_Status Anterior',
                                'Status_Re/Fi':'Status_Re/Fi Anterior',
                                'Status Vencimento':'Status/Venc anterior'}, inplace=True)       
    df_base = Pd.merge(df_base, df_anterior, on = 'FI_Code', how = 'left')  
    
    #Cria na Planilha Base as colunas Dia até Atraso, Status de Vencimento, Equipe de Atendimento, Tempo para conclusão e subir a evidência, Outras colunas
    df_base['Current due Date'] = Pd.to_datetime(df_base['Current due Date'], errors='coerce')
    df_base['Dias Até Atraso'] = (df_base['Current due Date'] - dta_ref).dt.days
    df_base['Tempo para conclusão'] = (df_base['Closing Date'] - df_base['Issue Date']).dt.days
    df_base['Tempo para subir evidencia'] = (df_base['Current due Date'] - df_base['Submit Date']).dt.days
    df_base['Status Vencimento'] = df_base.apply(lambda row:Status_Vencimento(row['Dias Até Atraso'], row['Status_Re/Fi'], row['Re_Status']), axis=1)
    df_base['Area de Atendimento'] = df_base.apply(lambda row:area_responsavel(row['Manager'], row['Segmento Brasil']), axis=1)
    df_base['Recalendarizada'] = df_base.apply(lambda row:Recalendarizacao(row['Data de Vencimento Anterior'], row['Current due Date'], row['Initial Due Date']), axis=1)
    df_base['Alterações do fechamento'] = df_base.apply(lambda row:AlteracoesFechamento(row['Re_Status'], row['Re_Status Anterior'],row['Status_Re/Fi Anterior'],row['Status_Re/Fi']), axis=1)
    
    #Organiza as colunas de forma que facilite a análise e define uma cor para grupos de colunas especificos
    colunas_relevantes={"002060":['Code','FI_Code','MO_Code','Short Name','Use Code','Tier','Severity','Description'],
                        "0F9ED5":['Status Vencimento','Status/Venc anterior','Re_Status','Re_Status Anterior','Alterações do fechamento','Fi_Status','Status_Re/Fi','Status_Re/Fi Anterior'],
                        "FF0000":['Dias Até Atraso','Current due Date','Data de Vencimento Anterior','Initial Due Date','Recalendarizada',
                                 'Qntd de Recalendarizações','Issue Date','Sign Off Date','Closing Date','Delete Date'],
                        "E97132":['Responsible','Manager','Responsible Legal Entity','Area de Atendimento'],
                        "A02B93":['Segmento Brasil','Scope','Subscope','Type','Subtype','Type of Regulatory'],
                        "4EA72E":['Team','Team Comments','Team Name','Edges','Final Rating'],
                        "262626":['Submit Date','Submit Comments','Review Date','Review Comments','Tempo para subir evidencia','Tempo para conclusão']        
    }
    
    colunas_iniciais=[]
    for c in colunas_relevantes.values():
        for x in c:
            colunas_iniciais.append(x)
    
    restante = [c for c in df_base.columns if c not in colunas_iniciais]
    df_base = df_base[colunas_iniciais + restante]
    df_base.pop('Unit Affected (Informe CRO)')
    
    #Passa o Data Frame criado para uma Planilha no Excel
    wb = op.load_workbook(r"template.xlsx")
    mes = datetime.now().strftime("%B").capitalize()
    wb.worksheets[0].title = f'Base_Monet_{mes}'
    e = wb[f'Base_Monet_{mes}']
    
    for num_linha, linha in enumerate(dataframe_to_rows(df_base, index=False,header=True),1):
        for num_coluna, valor in enumerate(linha,1):
            celula = e.cell(row= num_linha, column=num_coluna, value=valor)
            if num_linha == 1:
                celula.font = op.styles.Font(bold=True)
                celula.alignment = op.styles.Alignment(horizontal='center', vertical='center')
                celula.fill= op.styles.PatternFill(fill_type='solid', fgColor='A6A6A6')
                
                for cor,col in colunas_relevantes.items():
                    if celula.value in col: 
                        celula.fill= op.styles.PatternFill(fill_type='solid', fgColor=cor)
                        celula.font = op.styles.Font(color='FFFFFF', bold=True) 
    
    colunas_com_datas = ["S","T","U","X","Y","Z","AA","AB","AC","AV","AX","BG","BV"]       
    largura_col(e)
    for i in colunas_com_datas:
             Ajuste_data(i, e)

    nome_arq= f'Recos_fechamento_teste.xlsx'
    wb.save(local+nome_arq)
    Overview_Recos.Overview_forun(local+nome_arq, df_base)

    return df_base 

